"""
This connection opens an excel file and reads
"""

import os
import json
from shutil import copyfile
from openpyxl import Workbook, load_workbook
from common.bin.root_folder import aleph_root_folder


class ExcelConnection:

    def __init__(self, file, read_function):
        self.file = file
        self.connected = False
        self.read_function = read_function

        # Optional
        self.sheets_to_read = []
        self.watch_complete_file = False
        self.read_from_copy = False
        self.add_rows_only_when_complete = False
        self.max_editing_rows = 100
        self.include_past_records = False
        self.nullify_error_cells = True  # Changes error values (i.e. '#DIV/0!') to None

        # Aux
        self.temp_file = os.path.join(aleph_root_folder, "local", "temp", os.path.basename(self.file))
        self.last_row_file = self.temp_file + ".dat"
        self.last_row = {}
        self.last_modified = 0
        self.editing_rows = {}

    def connect(self):
        # Check if the excel file exists
        if os.path.isfile(self.file):
            self.connected = True
        else:
            raise Exception("File does not exist")

        # Check if we have a .dat file that contains the last scanned row
        if os.path.isfile(self.last_row_file):
            try:
                f = open(self.last_row_file)
                content = json.loads(f.read())
                f.close()
                self.last_row = content["last_row"]
                self.last_modified = content["last_modified"]
            except:
                pass

    def do(self):
        import time

        if not self.connected: raise Exception("Not connected")

        # Get the last modified timestamp from file
        # If file hasn't changed, do nothing
        last_modified = os.path.getmtime(self.file)
        if not last_modified > self.last_modified: return {}

        self.last_modified = last_modified

        # Copy the file because if it might be locked if open
        if self.read_from_copy:
            copyfile(self.file, self.temp_file)
            workbook = load_workbook(self.temp_file, data_only=True)
        else:
            workbook = load_workbook(self.file, data_only=True)

        # Get all sheets
        sheets_to_read = self.sheets_to_read
        if len(sheets_to_read) == 0:
            sheets_to_read = workbook.sheetnames

        # Main loop ---
        data = {}

        # For each sheet
        for sheet in sheets_to_read:
            data[sheet] = []

            # Get last column and last row
            last_col = workbook[sheet].max_column + 1
            last_row = workbook[sheet].max_row+1

            # Check if last row is known (if not it's the first run ever)
            if sheet not in self.last_row:
                self.editing_rows[sheet] = {}

                # Get last row (we do it this way to avoid problems with workbook[sheet].max_row,
                # because sometimes it may be greater than what it should be)
                r = 1
                while workbook[sheet].cell(r, 1).value is not None or workbook[sheet].cell(r, 2).value is not None or workbook[sheet].cell(r+1, 1).value is not None or workbook[sheet].cell(r+1, 2).value is not None:
                    r += 1
                self.last_row[sheet] = r

                # Include past records or add records to editing rows on first run
                if self.include_past_records or self.watch_complete_file:
                    for r in range(1, last_row):

                        # Get row as list
                        row_data = []
                        for c in range(1, last_col):
                            cell_value = workbook[sheet].cell(r, c).value
                            if self.nullify_error_cells and isinstance(cell_value, str) and cell_value.startswith("#") and cell_value.endswith("!"): cell_value = None
                            row_data.append(cell_value)

                        # Add row to data if include past records
                        if self.include_past_records:
                            data[sheet].append(row_data)

                        # Also add row to editing rows if watching all file
                        if self.watch_complete_file:
                            self.editing_rows[sheet][r] = row_data

                # Continue to next sheet
                continue

            # Detect row changes on editing rows
            if sheet in self.editing_rows:

                # For row in editing rows, see if there has been any change
                for r in self.editing_rows[sheet]:

                    # Get row as list
                    row_data = []
                    for c in range(1, last_col):
                        cell_value = workbook[sheet].cell(r, c).value
                        if self.nullify_error_cells and isinstance(cell_value, str) and cell_value.startswith("#") and cell_value.endswith("!"): cell_value = None
                        row_data.append(cell_value)

                    # If row length has changed, then a column has been added or removed
                    # Then clear all the editing rows
                    if len(row_data) != len(self.editing_rows[sheet][r]):
                        self.editing_rows[sheet] = {}
                        break

                    # If row values changed
                    if False in [self.editing_rows[sheet][r][i] == row_data[i] for i in range(0, len(row_data))]:
                        self.editing_rows[sheet][r] = row_data
                        # Add row to data
                        if not self.add_rows_only_when_complete or (self.add_rows_only_when_complete and None not in row_data):
                            data[sheet].append(row_data)

            # Fill editing rows if empty
            if self.watch_complete_file and (sheet not in self.editing_rows or len(self.editing_rows[sheet]) == 0):
                self.editing_rows[sheet] = {}

                for r in range(1, last_row):
                    row_data = []
                    for c in range(1, last_col):
                        cell_value = workbook[sheet].cell(r, c).value
                        if self.nullify_error_cells and isinstance(cell_value, str) and cell_value.startswith("#") and cell_value.endswith("!"): cell_value = None
                        row_data.append(cell_value)
                    self.editing_rows[sheet][r] = row_data

            # Scan for new rows
            r = self.last_row[sheet]
            while workbook[sheet].cell(r, 1).value is not None and workbook[sheet].cell(r, 2).value is not None:

                # Get row as list
                row_data = []
                for c in range(1, last_col):
                    cell_value = workbook[sheet].cell(r, c).value
                    if self.nullify_error_cells and isinstance(cell_value, str) and cell_value.startswith("#") and cell_value.endswith("!"): cell_value = None
                    row_data.append(cell_value)

                # Add new row to editing rows
                if sheet not in self.editing_rows:
                    self.editing_rows[sheet] = {}
                self.editing_rows[sheet][r] = row_data

                # Keep rows in edition length below max
                if not self.watch_complete_file and len(self.editing_rows[sheet]) > self.max_editing_rows:
                    del self.editing_rows[sheet][min(self.editing_rows[sheet].keys())]

                # Add to data if complete
                if not self.add_rows_only_when_complete or (self.add_rows_only_when_complete and None not in row_data):
                    data[sheet].append(row_data)

                r += 1

            # Update last row
            self.last_row[sheet] = r

        # -------------

        # Close workbook
        workbook.close()

        # Save last row to file
        try:
            f = open(self.last_row_file, "w+")
            f.write(json.dumps({"last_row": self.last_row, "last_modified": self.last_modified}))
            f.close()
        except:
            pass

        # Pass data to read_function
        return self.read_function(data)
