"""

"""

from ....connections.connection import Connection
from ....common.file_handler import FileHandler
from openpyxl import load_workbook

COLUMN_LETTERS = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"
COLUMN_LETTERS = [x for x in COLUMN_LETTERS]
COLUMN_LETTERS += ["A" + x for x in COLUMN_LETTERS]


class ExcelConnection(Connection):

    def __init__(self, client_id=""):
        super().__init__(client_id)
        self.report_by_exception = True
        self.clean_on_read = False

        # File path to the Excel
        self.file = ""

        # Optional
        self.read_from_copy = True                  # Copy file before reading
        self.number_of_watching_rows = 100          # Watch changes on the last number_of_watching_rows rows
        self.columns = COLUMN_LETTERS               # List of the names of the columns
        self.include_row_number = True              # Include row number in returned records
        self.temp_folder = ""

        # Private
        self.file_handler = None

    # ===================================================================================
    # Open
    # ===================================================================================
    def open(self):
        self.file_handler = FileHandler(self.file, self.read_from_copy)
        self.file_handler.temp_folder = self.temp_folder

    def close(self):
        self.file_handler = None

    # ===================================================================================
    # Read
    # ===================================================================================
    def read(self, key, **kwargs):
        # If data hasn't changed, don't do anything
        if not self.file_handler.file_has_been_modified(key): return []

        # Open workbook
        file = self.file_handler.get_file_for_reading(key)
        workbook = load_workbook(file, data_only=True)
        sheet = workbook[key]

        # Get last column
        last_col = sheet.max_column + 1

        # Get last row
        r = 1
        while not self.__following_cells_are_empty__(sheet, r): r += 1
        last_row = r

        # Get begin row
        begin_row = 1
        if last_row > self.number_of_watching_rows:
            begin_row = last_row - self.number_of_watching_rows

        data = []
        # For each row in range
        for r in range(begin_row, last_row):

            # Add row number
            row_data = {"id_": str(int(self.file_handler.file_creation_timestamp)) + str(r)}
            if self.include_row_number: row_data["n"] = r

            # For each column, get cell value
            for c in range(1, last_col):
                cell_value = sheet.cell(r, c).value
                if self.__cell_value_is_error__(cell_value): continue
                row_data[COLUMN_LETTERS[c-1]] = cell_value

            # Add data
            data.append(row_data)

        # Return
        return data

    # ===================================================================================
    # Aux
    # ===================================================================================
    @staticmethod
    def __following_cells_are_empty__(sheet, start_row):
        for i in range(0, 5):
            for j in range(1, 10 + 1):
                v = sheet.cell(start_row + i, j).value
                if v is not None:
                    return False

        return True

    @staticmethod
    def __cell_value_is_error__(cell_value):
        if isinstance(cell_value, str) and cell_value.startswith("#") and cell_value.endswith("!"):
            return True
        return False
