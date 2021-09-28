import datetime
import openpyxl

from django.http import HttpResponse, JsonResponse


# ===================================================================
# JSON
# ===================================================================

def json_response(dictionary):
    return JsonResponse(dictionary, safe=False)


# ===================================================================
# Excel
# ===================================================================

def xlsx_response(wb, filename):
    """

    """
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=' + filename + '.xlsx'
    wb.save(response)
    return response


def excel_auto_merge_cells(row1, row2, ws):
    #
    #

    column_names = ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O", "P", "Q", "R", "S", "T",
                    "U", "V", "W", "X", "Y", "Z"]
    column_names = column_names + ["A" + c for c in column_names]

    m = ""
    mi = 0
    for i in range(0, len(row1)):
        if row1[i] == row2[i] or row2[i] == "":
            ws.merge_cells(column_names[i] + '1:' + column_names[i] + '2')
        if row1[i] != m:
            if i - mi > 1:
                ws.merge_cells(column_names[mi] + '1:' + column_names[i - 1] + '1')
            m = row1[i]
            mi = i


def excel_prettify(ws, headers=1):
    openpyxl.worksheet.dimensions.ColumnDimension(ws, bestFit=True)

    column_names = tuple(openpyxl.utils.get_column_letter(col_number + 1) for col_number in range(ws.max_column))
    for column in ws.columns:
        # Auto fit column widths
        # ws.column_dimensions[column].bestFit = True
        as_text = lambda v: str(v) if v is not None else ""
        length = max(len(as_text(cell.value)) for cell in column)
        ws.column_dimensions[column[0].column_letter].width = length * 1.15

        # Make headers bold and centered
        for i in range(0, headers):
            ws[column[0].column_letter + str(i + 1)].font = openpyxl.styles.Font(bold=True)
            ws[column[0].column_letter + str(i + 1)].alignment = openpyxl.styles.alignment.Alignment(horizontal="center")


# ===================================================================
# PDF
# ===================================================================
def pdf_response(template, extras, request, filename):
    from django.template.loader import render_to_string
    from weasyprint import HTML, CSS

    html_string = render_to_string(template, extras)
    html = HTML(string=html_string, base_url=request.build_absolute_uri())
    pdf_file = html.write_pdf()
    response = HttpResponse(pdf_file, content_type='application/pdf')
    response['Content-Disposition'] = f'filename="{filename}.pdf"'
    return response


# ===================================================================
# Other
# ===================================================================

def duration_hhmmss(t1, t2):
    return str(datetime.datetime.strptime(t2, "%H:%M:%S") - datetime.datetime.strptime(t1, "%H:%M:%S"))


def nest_dictionary(source):
    target = {}
    for d in source:
        leaf = d
        subtarget = target
        while "." in leaf:
            root = leaf[0:leaf.index(".")]
            leaf = leaf[leaf.index(".")+1:]
            if root not in target: target[root] = {}
            subtarget = target[root]
        subtarget[leaf] = source[d]
    return target
